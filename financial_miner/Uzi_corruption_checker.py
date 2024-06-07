import numpy as np
import pandas as pd
from datetime import date, timedelta
from pylab import plt, mpl
import yfinance as yf
from openpyxl import Workbook, load_workbook
import os
import datetime
from tqdm import tqdm






industry_tickers = pd.read_excel("C:/Users/Nishesh Shukla\VScode\My Main projects\My projects on trade\Quants\Industries_tickers.xlsx")













root = "H:\My Drive"
data_storage = "DATA"


path = root+data_storage
path = os.path.join(root, data_storage) 

if not os.path.isdir(path):
    os.makedirs(path)
    
def directory_manager(industry,interval):
    
    interval_dictionary_two = {
        '1m':'One_min',
        '2m':'Two_min',
        '5m':'Five_min'
    }
    
    tick_path = os.path.join(path , interval_dictionary_two[interval] , industry)
    
    if not os.path.isdir(tick_path):
        os.makedirs(tick_path)

    return tick_path




















def excel_appending(industry_path,ticker,interval,industry):
    tick_path = industry_path + "/" + ticker + '.xlsx'
    
            
    try :
        wb = load_workbook(tick_path)
        ws1 = wb.active
        wb.save(tick_path)
        
    except:
        jury(ticker,interval,industry,tick_path)
    
    



def jury(ticker,interval,industry,tick_path):
    jury_path = "C:/Users/Nishesh Shukla/VScode/My Main projects/My projects on trade/Quants/jury.xlsx"
    if not os.path.isfile(jury_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.append(["date","ticker","interval","industry","tick_path"])
        ws1.append([str(date.today()),ticker,interval,industry,tick_path])
        
        wb.save(jury_path)
    else:
        wb = load_workbook(jury_path)
        ws1 = wb.active
        ws1.append([str(date.today()),ticker,interval,industry,tick_path])
        
        wb.save(jury_path)




















name = list(industry_tickers.columns)
intervals = ['1m','2m','5m']

            # for appending new row ----> of no use right now
for interval in intervals:
    for industry in tqdm(name):
        industry_path = directory_manager(industry, interval)
        for ticker in industry_tickers[industry].dropna(axis=0):
            excel_appending(industry_path,ticker+'.NS',interval,industry)









