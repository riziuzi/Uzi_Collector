import pandas as pd
from datetime import date, timedelta
import yfinance as yf
from openpyxl import Workbook, load_workbook
import os, sys
import datetime
from tqdm import tqdm
from Layla_My_Bot import Layla_message 
from Uzi_configuration import *












#                                                                   Directory Management

    
def directory_manager(industry,interval,root):
    
    data_storage = "DATA"
    # path = root+data_storage
    path = os.path.join(root, data_storage) 

    if not os.path.isdir(path):
        os.makedirs(path)
        
    interval_dictionary_two = {
        '1m':'One_min',
        '2m':'Two_min',
        '5m':'Five_min'
    }
    
    tick_path = os.path.join(path , interval_dictionary_two[interval] , industry)
    
    if not os.path.isdir(tick_path):
        os.makedirs(tick_path)

    return tick_path











#                                                                       Excel Management


def excel_appending(root,industry_path,ticker,interval,industry,empty_check=False):
    new_flag=False
    tick_path = industry_path + "/" + ticker + '.xlsx'
    
    
    
    if not os.path.isfile(tick_path):
        wb = Workbook()
        ws1 = wb.active
        new_flag=True
        try:
            raw = downloader(ticker, interval, new=True)
            ws1.append(list(raw.columns))
    
        except:
            print("No1")
            raw = pd.DataFrame()
            
        
    else:        
#         if empty_check == True:
#             return       
#         else:
        try:
            wb = load_workbook(tick_path)
            ws1 = wb.active
            last_date = last_date_checker(tick_path)
            print(last_date)
            try:
                raw = downloader(ticker, interval, last_date= last_date)
            except:
                print("No2")
                raw = pd.DataFrame()
        except:
            jury(ticker,interval,industry,tick_path,root=root)
            return
            
    
    if (not bool(len(raw.index))) & b_day(date.today()) or (not bool(len(raw.index)))&new_flag:
        keeper(ticker,interval,industry,root)
    
    for index,row in raw.iterrows(): 
         ws1.append(row.to_list())
    
    wb.save(tick_path)
    wb.close()












#                                                          Downloader, LAST_DATE_CHECKER & B-Checker


def b_day(date):
    return bool(len(pd.bdate_range(date,date)))

def last_date_checker(string):
    df = pd.read_excel(string)
    if df.empty:
        return None
    else:
        return datetime.datetime.strptime(df[df.columns[0]][len(df[df.columns[0]])-1][0:10], '%Y-%m-%d')

def downloader(ticker,interval,new= False, last_date=None):
    
    interval_dictionary = {
        '1m':None,
        '2m':str(date.today()-timedelta(59)),                                       # -----> I DIDN'T APPLIED STRING CONVERSION OF DATETYPE (now resolved)!!!!!! LOL 
        '5m':str(date.today()-timedelta(59))
    }
    
    if new:
        raw = yf.download(ticker,
                          interval=interval,
                          start=interval_dictionary[interval]
                         )
    elif last_date != None:
        raw = yf.download(ticker,
                        interval=interval,
                        start=last_date + timedelta(1)
                        )
    else:
        raw = yf.download(ticker,
                        interval=interval,
                        start=interval_dictionary[interval]
                        )
        
    raw = raw.reset_index()
    raw[raw.columns[0]]=raw[raw.columns[0]].astype(str)
    return raw
    









#                                                   Keeper!


def keeper(ticker,interval,industry,root):
    keeper_path = root+"/keeper.xlsx"
    if not os.path.isfile(keeper_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.append(["date","ticker","interval","industry"])
        ws1.append([str(date.today()),ticker,interval,industry])
        
        wb.save(keeper_path)
        wb.close()
        
    else:
        wb = load_workbook(keeper_path)
        ws1 = wb.active
        ws1.append([str(date.today()),ticker,interval,industry])
        
        wb.save(keeper_path)
        wb.close()
        










#                                                   Jury!



def jury(ticker,interval,industry,tick_path,root):
    jury_path = root+"/jury.xlsx"
    if not os.path.isfile(jury_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.append(["date","ticker","interval","industry","tick_path"])
        ws1.append([str(date.today()),ticker,interval,industry,tick_path])
        
        wb.save(jury_path)
        wb.close()
        
    else:
        wb = load_workbook(jury_path)
        ws1 = wb.active
        ws1.append([str(date.today()),ticker,interval,industry,tick_path])
        
        wb.save(jury_path)
        wb.close()
        





















#                                                        Initiator, Bussiness_Day Checker


def collector(industry_tickers=pd.read_excel(os.path.join(control,"Industries_tickers.xlsx")),root="C:\My Drive"):
    name = list(industry_tickers.columns)
    intervals = ['1m','2m','5m']

    pbar = tqdm(total=len(name)*len(intervals))


    if (datetime.datetime.now().time()<datetime.datetime.strptime("9 00", "%H %M").time()) or (datetime.datetime.now().time()>datetime.datetime.strptime("16 00", "%H %M").time()):            # for appending new row ----> of no use right now
        for interval in intervals:
            
            for industry in name:
                industry_path = directory_manager(industry, interval,root)
                for ticker in industry_tickers[industry].dropna(axis=0):
                    blockPrint()
                    excel_appending(industry_path=industry_path,ticker=ticker+'.NS',interval=interval,industry=industry,root=root)
                    enablePrint()
                pbar.update(1)
            Layla_message(f'Industries for {interval} Completed!')
    else:
        print("Can't update files during market hours!")
        Layla_message("Can't Update files in market Hours!")
        return
            
    pbar.close()
    print("Finished!")
    print(f"keeper.xlsx & jury.xlsx file saved at '{root}'")
    print(f"Data stored at {root}\DATA")
    Layla_part(root=root)














#                                                                    OTHERS


# DisablePrinting
def blockPrint():
    sys.stdout = open(os.devnull, 'w')

# RestorePrinting
def enablePrint():
    sys.stdout = sys.__stdout__



def Layla_part(root):
    
    
    Layla_message("Finished!")
    
    
    
    df = pd.read_excel(root+"/keeper.xlsx")
    if len(df.index) > 10:
        Layla_message("Absentees are : ",len(df.index))
    else:
        Layla_message("Absentese are : ")
        for i in range(len(df.index)):
            Layla_message(df.iloc[[i]])
    
    
    
    try:
        Layla_message("& Corrupts are :")
        df = pd.read_excel(root+"/jury.xlsx")
        if len(df.index)>10:
            Layla_message(len(df.index))
        else:
            for i in range(len(df.index)):
                Layla_message(df.iloc[[i]])
    except:
        # Layla_message(" No jurry.xlsx found!")
        pass
        
























if __name__ == "__main__":
    Layla_message("Started Uzi_Collector!")
    collector(root=root)
    






# Future resolution: File appending (Failed !)


# FilePath = "your excel path"
# ExcelWorkbook = load_workbook(FilePath)
# writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
# writer.book = ExcelWorkbook
# df.to_excel(writer, sheet_name = 'your sheet name')
# writer.save()
# writer.close()                









